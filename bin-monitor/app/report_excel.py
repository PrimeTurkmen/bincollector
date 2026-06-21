"""Export the bin-collection report as a multi-sheet .xlsx (management-friendly)."""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from . import queries


def _sheet(wb, title, headers, rows):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E5F")
    for r in rows:
        ws.append(r)
    return ws


def build_workbook() -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    k = queries.kpis()
    s = wb.create_sheet("Summary")
    s.append(["AIMS Bin Collection — Compiled Report"])
    s["A1"].font = Font(bold=True, size=14)
    s.append([])
    for label, key in [("Valid lift cycles", "total_lifts"), ("Matched <=15m", "matched"),
                       ("Near 15-30m", "near1"), ("Near 30-50m", "near2"),
                       ("Unmatched >50m", "unmatched"), ("Expected bins", "bins")]:
        s.append([label, k[key]])

    vd = queries.vehicle_day_summary()
    _sheet(wb, "Vehicle Day Summary",
           ["AgentId", "Plate", "Day", "Valid Lifts", "Matched <=15m", "Unique Bins <=15m",
            "Near 15-30m", "Near 30-50m", "Unmatched >50m", "Avg Distance m"],
           [[r["agent_id"], r["plate"], str(r["day"]), r["valid_lifts"], r["matched"],
             r["unique_bins"], r["near1"], r["near2"], r["unmatched"], r["avg_distance_m"]] for r in vd])

    ar = queries.area_summary()
    _sheet(wb, "Area Summary",
           ["Area Name", "Expected Bins", "Nearby Lifts", "Matched <=15m", "Near 15-30m", "Near 30-50m"],
           [[r["area_name"], r["expected_bins"], r["nearby_lifts"], r["matched"], r["near1"], r["near2"]] for r in ar])

    lf = queries.lifts(limit=100000)
    _sheet(wb, "Lift Details",
           ["AgentId", "Plate", "Day", "LiftNo", "StartTime", "DurationSec", "Lat", "Lon",
            "Matched Bin", "Area", "Bin Size", "DistanceM", "Band", "SuggestedStatus",
            "StartAngle", "LastHighAngle", "ResetAngle"],
           [[r["agent_id"], r["plate"], str(r["day"]), r["lift_no"], r["start_time"], r["duration_s"],
             r["lat"], r["lon"], r["matched_bin_id"], r["area_name"], r["bin_size"], r["distance_m"],
             r["distance_band"], r["suggested_status"], r["start_angle"], r["last_high_angle"],
             r["reset_angle"]] for r in lf])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
