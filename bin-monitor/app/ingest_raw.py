"""Parse the Pilot sensor-tracing .xlsx exports into raw_points.

Used to seed the demo from the real 3-day trial data. The angle (Param463),
ignition (DIS1), movement and battery values are parsed out of the device's
'raw data' string, which is always present, with the dedicated columns as
fallback.
"""
import re
import glob
import os
import openpyxl

from . import db, config
from .params import param as _param


def _agent_id_from_name(path: str):
    m = re.search(r"Agent(\d+)", os.path.basename(path))
    return int(m.group(1)) if m else None


def ingest_file(path: str) -> int:
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    header = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = {h: i for i, h in enumerate(header)}
    agent_id = _agent_id_from_name(path)
    src = os.path.basename(path)

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        raw = r[idx.get("raw data")] if "raw data" in idx else None
        rows.append((
            agent_id,
            r[idx["ts"]],
            r[idx["latitude"]],
            r[idx["longitude"]],
            r[idx["speed"]],
            _param(raw, "Param463"),
            _param(raw, "DIS1"),
            _param(raw, "Mvmnt"),
            _param(raw, "Vsourse"),
            src,
        ))
    wb.close()

    with db.connect() as conn:
        with conn.cursor() as cur:
            cur.executemany(
                """INSERT INTO raw_points
                   (agent_id, ts, lat, lon, speed, angle, dis1, mvmnt, vsourse, source_file)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                   ON CONFLICT (agent_id, ts) DO NOTHING""",
                rows,
            )
        conn.execute(
            "INSERT INTO ingest_runs (kind, agent_id, rows, note) VALUES (%s,%s,%s,%s)",
            ("rawfile", agent_id, len(rows), src),
        )
    return len(rows)


def ingest_all(directory: str = None) -> int:
    directory = directory or config.RAW_DATA_DIR
    files = sorted(glob.glob(os.path.join(directory, "*.xlsx")))
    total = 0
    for f in files:
        n = ingest_file(f)
        total += n
        print(f"  ingested {os.path.basename(f)}: {n} rows")
    print(f"raw_points ingested: {total} from {len(files)} files")
    return total
