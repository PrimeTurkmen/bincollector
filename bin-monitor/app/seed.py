"""Seed reference data: vehicles + geozones from Pilot, bin master from the report."""
import re
import json
import openpyxl

from . import db, config
from .pilot_client import PilotClient
from .geometry import polygon_center, haversine_m, BinIndex

PLATE_RE = re.compile(r"B\d{4,6}")


def _plate(object_name: str) -> str:
    m = PLATE_RE.search(object_name or "")
    return m.group(0) if m else (object_name or "")


def _parse_coord(s):
    if not s:
        return (None, None)
    parts = str(s).split(",")
    try:
        return (float(parts[0].strip()), float(parts[1].strip()))
    except (ValueError, IndexError):
        return (None, None)


def seed_vehicles(client: PilotClient) -> int:
    vs = client.vehicles()
    rows = []
    for v in vs:
        aid = v.get("agentid") or v.get("agent_id")
        if config.TRIAL_AGENT_IDS and aid not in config.TRIAL_AGENT_IDS:
            continue
        rows.append((aid, str(v.get("imei")), _plate(v.get("vehiclenumber")), v.get("vehiclenumber")))
    with db.connect() as conn:
        with conn.cursor() as cur:
            cur.executemany(
                """INSERT INTO vehicles (agent_id, imei, plate, object_name, has_angle_sensor)
                   VALUES (%s,%s,%s,%s,TRUE)
                   ON CONFLICT (agent_id) DO UPDATE SET imei=EXCLUDED.imei,
                     plate=EXCLUDED.plate, object_name=EXCLUDED.object_name""",
                rows,
            )
    print(f"vehicles seeded: {len(rows)}")
    return len(rows)


def seed_geozones(client: PilotClient) -> int:
    zones = client.geofences()
    rows = []
    for z in zones:
        pts = z.get("points") or []
        clat, clon = polygon_center(pts)
        rows.append((
            z["id"], z.get("name"), z.get("group_name"), z.get("type"),
            z.get("color"), json.dumps(pts), clat, clon,
        ))
    with db.connect() as conn:
        with conn.cursor() as cur:
            cur.executemany(
                """INSERT INTO geozones (pilot_id, name, group_name, type, color, points, center_lat, center_lon)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                   ON CONFLICT (pilot_id) DO UPDATE SET name=EXCLUDED.name,
                     group_name=EXCLUDED.group_name, points=EXCLUDED.points,
                     center_lat=EXCLUDED.center_lat, center_lon=EXCLUDED.center_lon""",
                rows,
            )
    print(f"geozones seeded: {len(rows)}")
    return len(rows)


def _find_header_row(ws, key="Unique ID", limit=8):
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=limit, values_only=True)):
        if any(str(c).strip() == key for c in row if c is not None):
            return i + 1  # 1-based row number
    return 4


def seed_bins(report_path: str = None) -> int:
    report_path = report_path or config.REPORT_XLSX
    wb = openpyxl.load_workbook(report_path, read_only=True)
    ws = wb["Bins Ditribution _Excel_"]
    hr = _find_header_row(ws)
    header = [str(c).strip() if c is not None else "" for c in
              next(ws.iter_rows(min_row=hr, max_row=hr, values_only=True))]
    idx = {h: i for i, h in enumerate(header)}

    def g(row, name):
        i = idx.get(name)
        return row[i] if i is not None and i < len(row) else None

    rows = []
    for r in ws.iter_rows(min_row=hr + 1, values_only=True):
        uid = g(r, "Unique ID")
        if not uid:
            continue
        lat, lon = _parse_coord(g(r, "Location"))
        rows.append((
            str(uid), g(r, "Asset Number"), g(r, "Area Name"), g(r, "Bin Size"),
            str(g(r, "Bin Quantity")) if g(r, "Bin Quantity") is not None else None,
            g(r, "Bin Placement"), g(r, "Bin Condition"), g(r, "Collection Vehicle Type"),
            lat, lon,
            g(r, "Nearest 1 Unique ID"), g(r, "Nearest 1 Distance (meters)"), g(r, "Nearest 1 Same Bin Size?"),
            g(r, "Nearest 2 Unique ID"), g(r, "Nearest 2 Distance (meters)"), g(r, "Nearest 2 Same Bin Size?"),
            g(r, "Nearest 3 Unique ID"), g(r, "Nearest 3 Distance (meters)"), g(r, "Nearest 3 Same Bin Size?"),
        ))
    wb.close()

    with db.connect() as conn:
        with conn.cursor() as cur:
            cur.executemany(
                """INSERT INTO bins (unique_id, asset_number, area_name, bin_size, quantity,
                       placement, condition, vehicle_type, lat, lon,
                       nearest1_id, nearest1_m, nearest1_same,
                       nearest2_id, nearest2_m, nearest2_same,
                       nearest3_id, nearest3_m, nearest3_same)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                   ON CONFLICT (unique_id) DO UPDATE SET area_name=EXCLUDED.area_name,
                       lat=EXCLUDED.lat, lon=EXCLUDED.lon""",
                rows,
            )
    print(f"bins seeded: {len(rows)}")
    return len(rows)


def link_geozones_to_bins() -> int:
    """Attach each bin-group geozone to its bin (by name match, else nearest centroid)."""
    bins = db.query("SELECT unique_id, lat, lon FROM bins WHERE lat IS NOT NULL")
    by_name = {b["unique_id"]: b for b in bins}
    index = BinIndex(bins)
    zones = db.query(
        "SELECT pilot_id, name, center_lat, center_lon FROM geozones WHERE group_name=%s",
        (config.BIN_GROUP_NAME,),
    )
    updates = []
    for z in zones:
        uid = None
        if z["name"] in by_name:
            uid = z["name"]
        elif z["center_lat"] is not None:
            b, d = index.nearest(z["center_lat"], z["center_lon"])
            if b and d is not None and d <= 25:
                uid = b["unique_id"]
        if uid:
            updates.append((uid, z["pilot_id"]))
    with db.connect() as conn:
        with conn.cursor() as cur:
            cur.executemany("UPDATE geozones SET bin_unique_id=%s WHERE pilot_id=%s", updates)
            cur.executemany("UPDATE bins SET geozone_id=%s WHERE unique_id=%s",
                            [(p, u) for u, p in updates])
    print(f"geozone<->bin links: {len(updates)}")
    return len(updates)
